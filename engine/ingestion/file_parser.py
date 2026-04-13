"""Parse multimodal files into a common text+metadata format.

Supports:
- PDF (via pdfplumber)
- Excel (.xlsx, .xls, .xlsm via openpyxl)
- Images (PNG, JPG — metadata only for now; OCR via OpenAI vision in Phase 3)
- Plain text (.txt, .md)

All parsers return a common shape::

    {
        "source_type": "file",
        "file_type": "pdf" | "excel" | "image" | "text",
        "filename": "...",
        "title": "...",            # filename without extension as default
        "content": "full text",    # concatenated across pages / sheets
        "metadata": {...},         # page count, sheet names, dimensions, etc.
    }
"""

from __future__ import annotations

import logging
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

TEXT_SUFFIXES = {".txt", ".md", ".log"}
PDF_SUFFIXES = {".pdf"}
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".csv"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif"}


@dataclass
class ParsedFile:
    source_type: str
    file_type: str
    filename: str
    title: str
    content: str
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# Individual parsers
# ---------------------------------------------------------------------------


def parse_pdf(path: Path) -> ParsedFile:
    import pdfplumber

    pages: list[str] = []
    page_count = 0
    with pdfplumber.open(path) as pdf:
        page_count = len(pdf.pages)
        for page in pdf.pages:
            text = page.extract_text() or ""
            pages.append(text.strip())
    content = "\n\n".join(p for p in pages if p)
    return ParsedFile(
        source_type="file",
        file_type="pdf",
        filename=path.name,
        title=path.stem.replace("_", " ").replace("-", " ").title(),
        content=content,
        metadata={"page_count": page_count, "char_count": len(content)},
    )


def _parse_legacy_xls(path: Path) -> ParsedFile:
    """Parse old .xls format via xlrd."""
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheets_data: list[dict[str, Any]] = []
    text_blocks: list[str] = []
    for sheet in book.sheets():
        rows: list[list[str]] = []
        for row_idx in range(sheet.nrows):
            cells = []
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                cells.append(str(value) if value is not None else "")
            if any(c.strip() for c in cells):
                rows.append(cells)
        sheets_data.append(
            {"name": sheet.name, "row_count": len(rows), "col_count": sheet.ncols}
        )
        if rows:
            header = f"=== {sheet.name} ==="
            body = "\n".join("\t".join(r) for r in rows[:200])
            text_blocks.append(f"{header}\n{body}")
    content = "\n\n".join(text_blocks)
    return ParsedFile(
        source_type="file",
        file_type="excel",
        filename=path.name,
        title=path.stem.replace("_", " ").replace("-", " ").title(),
        content=content,
        metadata={
            "sheets": [s["name"] for s in sheets_data],
            "sheet_details": sheets_data,
            "format": "xls",
        },
    )


def parse_excel(path: Path) -> ParsedFile:
    # CSV path — use the csv module so we don't pull pandas.
    if path.suffix.lower() == ".csv":
        import csv

        rows: list[list[str]] = []
        with path.open(encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append([str(c) for c in row])
        content_lines = ["\t".join(row) for row in rows]
        content = "\n".join(content_lines)
        return ParsedFile(
            source_type="file",
            file_type="excel",
            filename=path.name,
            title=path.stem.replace("_", " ").replace("-", " ").title(),
            content=content,
            metadata={"sheets": ["Sheet1"], "row_count": len(rows)},
        )

    # Legacy .xls format — openpyxl can't read this
    if path.suffix.lower() == ".xls":
        return _parse_legacy_xls(path)

    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets_data: list[dict[str, Any]] = []
    text_blocks: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
        sheets_data.append(
            {
                "name": sheet_name,
                "row_count": len(rows),
                "col_count": max((len(r) for r in rows), default=0),
            }
        )
        if rows:
            header = f"=== {sheet_name} ==="
            body = "\n".join("\t".join(r) for r in rows[:200])  # cap rows per sheet
            text_blocks.append(f"{header}\n{body}")
    wb.close()
    content = "\n\n".join(text_blocks)
    return ParsedFile(
        source_type="file",
        file_type="excel",
        filename=path.name,
        title=path.stem.replace("_", " ").replace("-", " ").title(),
        content=content,
        metadata={"sheets": [s["name"] for s in sheets_data], "sheet_details": sheets_data},
    )


def parse_image(path: Path) -> ParsedFile:
    from PIL import Image

    with Image.open(path) as img:
        size = img.size
        mode = img.mode
        fmt = img.format
    return ParsedFile(
        source_type="file",
        file_type="image",
        filename=path.name,
        title=path.stem.replace("_", " ").replace("-", " ").title(),
        content="",  # OCR happens downstream in the NLP layer
        metadata={
            "width": size[0],
            "height": size[1],
            "mode": mode,
            "format": fmt,
            "requires_ocr": True,
        },
    )


def parse_text(path: Path) -> ParsedFile:
    content = path.read_text(encoding="utf-8", errors="replace")
    return ParsedFile(
        source_type="file",
        file_type="text",
        filename=path.name,
        title=path.stem.replace("_", " ").replace("-", " ").title(),
        content=content,
        metadata={"char_count": len(content), "line_count": content.count("\n") + 1},
    )


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------


def parse(path: Path | str) -> ParsedFile:
    """Auto-detect file type and parse."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {p}")
    suffix = p.suffix.lower()
    if suffix in PDF_SUFFIXES:
        return parse_pdf(p)
    if suffix in EXCEL_SUFFIXES:
        return parse_excel(p)
    if suffix in IMAGE_SUFFIXES:
        return parse_image(p)
    if suffix in TEXT_SUFFIXES:
        return parse_text(p)
    raise ValueError(f"Unsupported file type: {suffix} ({p.name})")
