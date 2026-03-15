"""Multimodal file processors — PDF, image, audio, spreadsheet.

Per MASTER_BUILD_PLAN Phase 10:
- PDF (pdfplumber), image (Claude Vision), audio (Whisper), spreadsheet (openpyxl)
- Each processor returns extracted text + metadata
- Extracted data feeds into entity extraction → Jena ontology
"""

import io
from typing import Any

import structlog

from backend.core.config import settings

logger = structlog.get_logger()


class ProcessorResult:
    """Standard result from any file processor."""
    def __init__(
        self,
        text: str,
        metadata: dict[str, Any] | None = None,
        page_count: int | None = None,
        language: str | None = None,
        chunks: list[dict] | None = None,
    ):
        self.text = text
        self.metadata = metadata or {}
        self.page_count = page_count
        self.language = language
        self.chunks = chunks or []


def detect_processor(content_type: str, filename: str) -> str | None:
    """Determine which processor to use based on content type."""
    ct = content_type.lower()
    fn = filename.lower()

    if ct == "application/pdf" or fn.endswith(".pdf"):
        return "pdf"
    elif ct.startswith("image/") or fn.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
        return "image"
    elif ct.startswith("audio/") or fn.endswith((".mp3", ".wav", ".m4a", ".ogg", ".flac")):
        return "audio"
    elif ct in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
    ) or fn.endswith((".xlsx", ".xls", ".csv")):
        return "spreadsheet"
    elif ct.startswith("text/") or fn.endswith((".txt", ".md", ".json")):
        return "text"
    return None


async def process_pdf(file_data: bytes) -> ProcessorResult:
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber

        chunks = []
        full_text_parts = []
        metadata: dict[str, Any] = {}

        with pdfplumber.open(io.BytesIO(file_data)) as pdf:
            metadata["page_count"] = len(pdf.pages)
            metadata["pdf_metadata"] = pdf.metadata or {}

            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                full_text_parts.append(text)
                if text.strip():
                    chunks.append({
                        "content": text,
                        "page_number": i + 1,
                        "chunk_index": len(chunks),
                    })

                # Extract tables as text
                tables = page.extract_tables()
                for table in tables:
                    table_text = "\n".join(
                        " | ".join(str(cell or "") for cell in row)
                        for row in table
                    )
                    if table_text.strip():
                        chunks.append({
                            "content": f"[Table on page {i + 1}]\n{table_text}",
                            "page_number": i + 1,
                            "chunk_index": len(chunks),
                        })

        return ProcessorResult(
            text="\n\n".join(full_text_parts),
            metadata=metadata,
            page_count=metadata.get("page_count"),
            chunks=chunks,
        )
    except ImportError:
        logger.warning("pdfplumber_not_installed")
        return ProcessorResult(text="", metadata={"error": "pdfplumber not installed"})
    except Exception as e:
        logger.error("pdf_processing_failed", error=str(e))
        return ProcessorResult(text="", metadata={"error": str(e)})


async def process_image(file_data: bytes, filename: str) -> ProcessorResult:
    """Extract text/description from image using Claude Vision."""
    if not settings.ANTHROPIC_API_KEY:
        return ProcessorResult(text="", metadata={"error": "ANTHROPIC_API_KEY not configured"})

    try:
        import anthropic
        import base64

        client = anthropic.AsyncAnthropic(api_key=settings.ANTHROPIC_API_KEY)

        # Determine media type
        ext = filename.rsplit(".", 1)[-1].lower()
        media_types = {
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "gif": "image/gif",
            "webp": "image/webp",
        }
        media_type = media_types.get(ext, "image/jpeg")

        b64_data = base64.standard_b64encode(file_data).decode("utf-8")

        response = await client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": "Extract all text visible in this image. If it contains charts, tables, or diagrams, describe their content and data. Focus on any ESG-related information (environmental, social, governance). Return the extracted content as structured text.",
                    },
                ],
            }],
        )

        text = response.content[0].text
        return ProcessorResult(
            text=text,
            metadata={"source": "claude_vision", "media_type": media_type},
            chunks=[{"content": text, "chunk_index": 0}],
        )
    except Exception as e:
        logger.error("image_processing_failed", error=str(e))
        return ProcessorResult(text="", metadata={"error": str(e)})


async def process_audio(file_data: bytes, filename: str) -> ProcessorResult:
    """Transcribe audio using OpenAI Whisper API."""
    if not settings.OPENAI_API_KEY:
        return ProcessorResult(text="", metadata={"error": "OPENAI_API_KEY not configured for Whisper"})

    try:
        import httpx

        async with httpx.AsyncClient(timeout=120.0) as client:
            ext = filename.rsplit(".", 1)[-1].lower()
            response = await client.post(
                "https://api.openai.com/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {settings.OPENAI_API_KEY}"},
                files={"file": (filename, file_data, f"audio/{ext}")},
                data={"model": "whisper-1", "response_format": "verbose_json"},
            )
            response.raise_for_status()
            result = response.json()

        text = result.get("text", "")
        segments = result.get("segments", [])

        chunks = []
        # Group segments into ~500 char chunks
        current_chunk = ""
        for seg in segments:
            seg_text = seg.get("text", "")
            if len(current_chunk) + len(seg_text) > 500:
                if current_chunk:
                    chunks.append({
                        "content": current_chunk.strip(),
                        "chunk_index": len(chunks),
                    })
                current_chunk = seg_text
            else:
                current_chunk += " " + seg_text
        if current_chunk.strip():
            chunks.append({"content": current_chunk.strip(), "chunk_index": len(chunks)})

        return ProcessorResult(
            text=text,
            metadata={
                "source": "whisper",
                "language": result.get("language"),
                "duration": result.get("duration"),
            },
            language=result.get("language"),
            chunks=chunks,
        )
    except Exception as e:
        logger.error("audio_processing_failed", error=str(e))
        return ProcessorResult(text="", metadata={"error": str(e)})


async def process_spreadsheet(file_data: bytes, filename: str) -> ProcessorResult:
    """Extract data from Excel/CSV files."""
    try:
        ext = filename.rsplit(".", 1)[-1].lower()
        chunks = []
        all_text_parts = []
        metadata: dict[str, Any] = {}

        if ext == "csv":
            import csv
            reader = csv.reader(io.StringIO(file_data.decode("utf-8", errors="replace")))
            rows = list(reader)
            metadata["row_count"] = len(rows)
            metadata["columns"] = rows[0] if rows else []
            text = "\n".join(" | ".join(row) for row in rows)
            all_text_parts.append(text)
            # Chunk by ~50 rows
            for i in range(0, len(rows), 50):
                chunk_rows = rows[i:i + 50]
                chunk_text = "\n".join(" | ".join(row) for row in chunk_rows)
                chunks.append({"content": chunk_text, "chunk_index": len(chunks)})
        else:
            # Excel
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            metadata["sheet_names"] = wb.sheetnames
            metadata["sheet_count"] = len(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_text_parts = [f"[Sheet: {sheet_name}]"]
                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    rows_data.append(row_text)
                    sheet_text_parts.append(row_text)

                sheet_text = "\n".join(sheet_text_parts)
                all_text_parts.append(sheet_text)

                # Chunk by ~50 rows per sheet
                for i in range(0, len(rows_data), 50):
                    chunk_rows = rows_data[i:i + 50]
                    chunk_text = f"[Sheet: {sheet_name}]\n" + "\n".join(chunk_rows)
                    chunks.append({"content": chunk_text, "chunk_index": len(chunks)})

            wb.close()

        return ProcessorResult(
            text="\n\n".join(all_text_parts),
            metadata=metadata,
            chunks=chunks,
        )
    except ImportError as e:
        logger.warning("spreadsheet_processor_missing_dep", error=str(e))
        return ProcessorResult(text="", metadata={"error": f"Missing dependency: {e}"})
    except Exception as e:
        logger.error("spreadsheet_processing_failed", error=str(e))
        return ProcessorResult(text="", metadata={"error": str(e)})


async def process_text(file_data: bytes) -> ProcessorResult:
    """Process plain text files."""
    text = file_data.decode("utf-8", errors="replace")
    chunks = []
    # Chunk by ~1000 chars
    for i in range(0, len(text), 1000):
        chunk = text[i:i + 1000]
        if chunk.strip():
            chunks.append({"content": chunk, "chunk_index": len(chunks)})

    return ProcessorResult(text=text, chunks=chunks)


# Processor registry
PROCESSORS = {
    "pdf": process_pdf,
    "image": process_image,
    "audio": process_audio,
    "spreadsheet": process_spreadsheet,
    "text": process_text,
}


async def process_file(
    file_data: bytes,
    filename: str,
    processor_type: str,
) -> ProcessorResult:
    """Route to the appropriate processor."""
    processor_fn = PROCESSORS.get(processor_type)
    if not processor_fn:
        return ProcessorResult(text="", metadata={"error": f"Unknown processor: {processor_type}"})

    if processor_type in ("image", "audio", "spreadsheet"):
        return await processor_fn(file_data, filename)
    else:
        return await processor_fn(file_data)
